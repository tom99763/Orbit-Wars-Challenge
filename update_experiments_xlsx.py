"""Build / refresh experiments/experiments_log.xlsx.

One row per training experiment (bc_v1, bc_v2, ppo_v1, alphazero_v1,
impala_v1, …). Run this script any time an experiment finishes — it
overwrites the xlsx with the latest snapshot of the EXPERIMENTS list
below.

Add new experiments by appending to EXPERIMENTS. Each row is a dict
covering method summary, hyper-parameters, final results, eval win
rates, and notes.
"""

from __future__ import annotations

import pathlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


REPO = pathlib.Path(__file__).parent
OUT = REPO / "experiments" / "experiments_log.xlsx"


# ---------------------------------------------------------------
# Experiment records — EDIT THIS when a run finishes
# ---------------------------------------------------------------

EXPERIMENTS: list[dict] = [
    {
        "name": "bc_v1",
        "date": "2026-04-19",
        "type": "BC",
        "method": "Supervised classification of (target_planet, ships_bucket) on top-10 winner trajectories",
        "data": "~4k samples (41 winner trajectories from 50 scraped top-10 replays)",
        "model": "OrbitAgent 0.9M params, Set-Transformer 4 layers, d=128",
        "hparams": "15 epochs (SIGPIPE-killed at 30), lr=3e-4 cosine, batch=64",
        "final_loss": 2.32,
        "vs_random": 0.95,
        "vs_starter": 0.05,
        "vs_lb928": "n/a",
        "kaggle_score": "n/a",
        "status": "completed",
        "verdict": "Weak baseline; too little data",
        "notes": "Discovered SIGPIPE from tee+tail combo; moved to direct file writes",
    },
    {
        "name": "bc_v2",
        "date": "2026-04-19",
        "type": "BC",
        "method": "BC on mixed curriculum (500 starter games + 41 top-10 winners)",
        "data": "~230k samples, 830 winner trajectories",
        "model": "OrbitAgent 0.9M params",
        "hparams": "30 epochs, lr=3e-4 cosine, batch=128",
        "final_loss": 0.099,
        "vs_random": 0.95,
        "vs_starter": 0.20,
        "vs_lb928": 0.00,
        "kaggle_score": "ERROR (Validation Episode failed, submissionId 51835195)",
        "status": "completed + submitted",
        "verdict": "Best BC so far; submission errored on Kaggle grader (import > 2s overage?)",
        "notes": "Used as starting point for all subsequent RL runs",
    },
    {
        "name": "bc_v3",
        "date": "2026-04-19",
        "type": "BC",
        "method": "BC adding lb-928 rules-based planner games to curriculum",
        "data": "~331k samples, 1588 winner trajectories (720 lb-928 + 500 starter + 106 real)",
        "model": "OrbitAgent 0.9M params (same as bc_v2)",
        "hparams": "30 epochs, lr=3e-4 cosine, batch=128",
        "final_loss": 0.973,
        "vs_random": 0.95,
        "vs_starter": 0.05,
        "vs_lb928": 0.00,
        "kaggle_score": "n/a",
        "status": "completed",
        "verdict": "REGRESSED vs starter (20→5%). 0.9M model too small to absorb lb-928 complexity",
        "notes": "Key lesson: heterogeneous teachers dilute signal at fixed capacity",
    },
    {
        "name": "ppo_v1_from_bcv1",
        "date": "2026-04-19",
        "type": "PPO",
        "method": "PPO self-play starting from bc_v1 + teacher-KL(0.5→0.05) + shaped reward + PFSP pool",
        "data": "60 updates × 2 eps × varied game length",
        "model": "OrbitAgent 0.9M (+ value head active)",
        "hparams": "lr=2.5e-4, eps/update=2, 4p_prob=0.5, shape_weight=1",
        "final_loss": "n/a",
        "vs_random": "n/a",
        "vs_starter": 0.25,
        "vs_lb928": "n/a",
        "kaggle_score": "n/a",
        "status": "stopped (plateau 25% over 50 updates)",
        "verdict": "PPO can't escape weak BC starting point; teacher-KL drags toward it",
        "notes": "First PPO; highlighted need for stronger warmstart",
    },
    {
        "name": "ppo_v2_from_bcv2",
        "date": "2026-04-19",
        "type": "PPO",
        "method": "PPO from bc_v2 (stronger start) + teacher-KL 0.3 decaying",
        "data": "25 updates × 2 eps",
        "model": "OrbitAgent 0.9M",
        "hparams": "lr=2.5e-4, eps/update=2, 4p_prob=0.5",
        "final_loss": "n/a",
        "vs_random": "n/a",
        "vs_starter": 0.14,
        "vs_lb928": "n/a",
        "kaggle_score": "n/a",
        "status": "stopped (policy DEGRADING, 33% → 14%)",
        "verdict": "Small batch + teacher-KL incompatible; PPO setup broken",
        "notes": "Triggered redesign to A2C/IMPALA approach",
    },
    {
        "name": "alphazero_v1",
        "date": "2026-04-19",
        "type": "AlphaZero",
        "method": "Sampled MCTS self-play (K=4 sampled actions, n_sims=10/20) + policy-only deployment (mode A)",
        "data": "attempted 4-worker × 125 iters × 4 games = 500 games",
        "model": "OrbitAgent 0.9M",
        "hparams": "n_sims=10, k_samples=4, lr=1e-4, pol_coef=0.1",
        "final_loss": "iter 0 = 2.39 (only 1 iter completed before killed)",
        "vs_random": "n/a",
        "vs_starter": "n/a",
        "vs_lb928": "n/a",
        "kaggle_score": "n/a",
        "status": "killed (too slow; 7.6 min/iter → 16+ hr total)",
        "verdict": "MCTS per-turn cost too high for RTS with 1s actTimeout; env.deepcopy and jsonschema validation were bottlenecks",
        "notes": "Env-validation monkey-patch trick saved 22%; learned deepcopy speedup via steps-truncation",
    },
    {
        "name": "impala_v1",
        "date": "2026-04-19",
        "type": "IMPALA-lite (Lux-flavoured)",
        "method": "Sync A2C self-play with Orbit-Wars-specific reward shaping, decaying over 50 iters. No teacher, all seats contribute, advantage-normalised.",
        "data": "125 iters × 4 workers × 1 game each = 500 games (target)",
        "model": "OrbitAgent 0.9M",
        "hparams": "lr=1e-4, 4p_prob=0.5, shape_decay=50 iters, teacher_kl=0.0, grad_steps=2/iter",
        "final_loss": "RUNNING",
        "vs_random": "TBD",
        "vs_starter": "iter 9: 10% (10 games, high variance)",
        "vs_lb928": "iter 9: 0%",
        "kaggle_score": "TBD",
        "status": "running",
        "verdict": "TBD — in progress",
        "notes": "70-90s/iter, est 2-3 hours total. Eval watcher polling every 10 min.",
    },
]


# ---------------------------------------------------------------
# Workbook formatting
# ---------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="334488", end_color="334488", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_COLORS = {
    "running": "FFFBD1",
    "completed": "DFF7DF",
    "completed + submitted": "BEE7A5",
    "stopped (plateau 25% over 50 updates)": "F4D3D3",
    "stopped (policy DEGRADING, 33% → 14%)": "F4D3D3",
    "killed (too slow; 7.6 min/iter → 16+ hr total)": "E0D5F1",
}


def _write_experiments_sheet(wb: Workbook):
    ws = wb.active
    ws.title = "experiments"
    cols = [
        ("name", 22),
        ("date", 12),
        ("type", 22),
        ("method", 60),
        ("data", 50),
        ("model", 35),
        ("hparams", 55),
        ("final_loss", 16),
        ("vs_random", 11),
        ("vs_starter", 12),
        ("vs_lb928", 11),
        ("kaggle_score", 35),
        ("status", 35),
        ("verdict", 50),
        ("notes", 60),
    ]
    for col_idx, (name, width) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[c.column_letter].width = width

    for r, exp in enumerate(EXPERIMENTS, start=2):
        fill = PatternFill(start_color=STATUS_COLORS.get(exp.get("status"), "FFFFFF"),
                           end_color=STATUS_COLORS.get(exp.get("status"), "FFFFFF"),
                           fill_type="solid")
        for col_idx, (name, _) in enumerate(cols, start=1):
            val = exp.get(name, "")
            if isinstance(val, float):
                if name in ("vs_random", "vs_starter", "vs_lb928"):
                    val = f"{val:.0%}"
                else:
                    val = round(val, 4)
            c = ws.cell(row=r, column=col_idx, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.fill = fill
            c.border = BORDER

    ws.freeze_panes = "B2"   # freeze top row + first column


def _write_legend_sheet(wb: Workbook):
    ws = wb.create_sheet("legend")
    rows = [
        ("Column", "Meaning"),
        ("name", "Experiment tag (e.g., bc_v1, impala_v1)"),
        ("date", "Date of training run"),
        ("type", "Method family: BC / PPO / A2C / AlphaZero / IMPALA-lite"),
        ("method", "One-line description of approach"),
        ("data", "Training dataset composition + size"),
        ("model", "Architecture + params"),
        ("hparams", "Key hyperparameters"),
        ("final_loss", "Training loss at end (or 'RUNNING'/'n/a')"),
        ("vs_random", "Local win rate vs random agent (N=20)"),
        ("vs_starter", "Local win rate vs kaggle starter agent"),
        ("vs_lb928", "Local win rate vs lb-928 rules-based planner"),
        ("kaggle_score", "Public LB score if submitted"),
        ("status", "running / completed / stopped / killed"),
        ("verdict", "One-line takeaway"),
        ("notes", "Anything worth remembering"),
        ("", ""),
        ("Status colour code", ""),
        ("running", "pale yellow"),
        ("completed", "pale green"),
        ("completed + submitted", "bright green"),
        ("stopped", "pale red"),
        ("killed", "pale purple"),
    ]
    for r, (a, b) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=a).font = Font(bold=(r == 1 or r == 17))
        ws.cell(row=r, column=2, value=b)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60


def main() -> int:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _write_experiments_sheet(wb)
    _write_legend_sheet(wb)
    wb.save(OUT)
    print(f"wrote {OUT}  ({len(EXPERIMENTS)} experiments)")
    return 0


if __name__ == "__main__":
    import sys
    sys.exit(main())
